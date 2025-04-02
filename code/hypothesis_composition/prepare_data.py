import os
import shutil

def copy_folders_with_d0_json(distance_path, copy_path, name):
    for root, dirs, files in os.walk(distance_path):
        if 'merge.json' in files:
            relative_path = os.path.relpath(root, distance_path)
            target_folder = os.path.join(copy_path, name, relative_path)
            os.makedirs(target_folder, exist_ok=True)
            source_file = os.path.join(root, 'merge.json')
            target_file = os.path.join(target_folder, 'merge.json')
            shutil.copy2(source_file, target_file)
            print(f"Copied: {source_file} to {target_file}")

name = "Math"
distance_path = r"D:\\python\\spider\\distance\\" + name
copy_path = r"D:\\python\\generation"

copy_folders_with_d0_json(distance_path, copy_path, name)


import json
import pandas as pd
import re

name = "Chemistry"
json_path = r"D:\\python\\" + name + "\\result.json"
xlsx_path = r"D:\\python\\" + name + "\\result_with_overall.xlsx"

# Load JSON data
with open(json_path, 'r', encoding='utf-8') as f:
    json_data = json.load(f)

# Load Excel file
df = pd.read_excel(xlsx_path)

def clean_doi(doi):
    return re.sub(r'[\W_]+', '', doi.lower())

# Process DOI and update Note column
for index, row in df.iterrows():
    excel_doi = clean_doi(row['doi'])
    for item in json_data:
        json_doi = clean_doi(item['doi'])
        if excel_doi == json_doi:
            insp_values = [insp['insp'] for insp in item['inspiration']]
            note_content = ", ".join([f"insp{i+1}: {insp}" for i, insp in enumerate(insp_values)])
            df.at[index, 'Note'] = note_content
            break

# Save updated Excel file
df.to_excel(xlsx_path, index=False)
print("Processing completed!")


import pandas as pd
from openpyxl import load_workbook

def fix_excel_sheet(path):
    workbook = load_workbook(path)
    sheet_names = workbook.sheetnames
    if not sheet_names:
        raise ValueError("The Excel file contains no sheets.")
    
    original_sheet_name = sheet_names[0]
    original_sheet = workbook[original_sheet_name]
    
    if "Overall" in sheet_names:
        del workbook["Overall"]
    
    new_sheet = workbook.create_sheet(title="Overall")
    for row in original_sheet.iter_rows(values_only=True):
        new_sheet.append(row)
    
    del workbook[original_sheet_name]
    workbook.save(path)
    print(f"Successfully fixed the Excel file. New sheet 'Overall' created and old sheet '{original_sheet_name}' deleted.")

print(name)
path = r"D:\\python\\" + name + "\\result_with_overall.xlsx"
fix_excel_sheet(path)


import os
import shutil

def copy_xlsx_file(xlsx_path, copy_path):
    os.makedirs(copy_path, exist_ok=True)
    file_name = os.path.basename(xlsx_path)
    target_file = os.path.join(copy_path, file_name)
    shutil.copy2(xlsx_path, target_file)
    print(f"Copied: {xlsx_path} to {target_file}")

xlsx_path = r"D:\\python\\" + name + "\\result_with_overall.xlsx"
copy_path = r"D:\\python\\generation\\" + name

copy_xlsx_file(xlsx_path, copy_path)


import os
import json

def process_files(distance_path):
    for root, dirs, files in os.walk(distance_path):
        if 'd0.json' in files and '4o_retrieve_res.json' in files:
            d0_path = os.path.join(root, 'd0.json')
            retrieve_res_path = os.path.join(root, '4o_retrieve_res.json')
            true_retrieve_path = os.path.join(root, 'ture_retrieve.json')

            # Load d0.json and 4o_retrieve_res.json
            with open(d0_path, 'r', encoding='utf-8') as d0_file:
                d0_data = json.load(d0_file)
            with open(retrieve_res_path, 'r', encoding='utf-8') as retrieve_res_file:
                retrieve_res_data = json.load(retrieve_res_file)
                bkg_q = list(retrieve_res_data[0].keys())[0]

            # Create ture_retrieve.json data
            true_retrieve_data = [
                {bkg_q: [d0_data]},
                {bkg_q: [[1.0, 1.0]]}
            ]

            # Write to ture_retrieve.json
            with open(true_retrieve_path, 'w', encoding='utf-8') as true_retrieve_file:
                json.dump(true_retrieve_data, true_retrieve_file, ensure_ascii=False, indent=4)

            print(f"Generated: {true_retrieve_path}")

name_list = ["Cell Biology", "Chemistry", "Earth Science", "Material Science", "Physics", 
             "Energy Science", "Environmental Science", "Biology", "Business", "Law", 
             "Math", "Astronomy"]

for name in name_list:
    distance_path = r"D:\\python\\generation\\" + name
    process_files(distance_path)
